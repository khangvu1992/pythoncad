import win32com.client
import time
from openpyxl import Workbook



acad = win32com.client.Dispatch("AutoCAD.Application")
doc = acad.ActiveDocument

print("\n👉 Hãy chọn 1 object để lấy layer")

try:
    obj_pick = doc.Utility.GetEntity()[0]
except:
    print("❌ Hủy chọn")
    exit()


print("\nTYPE:", obj_pick.ObjectName)

for attr in dir(obj_pick):
    if attr.startswith("_"):
        continue

    try:
        value = getattr(obj_pick, attr)

        if callable(value):
            continue

        print(f"{attr:<30} = {value} ({type(value).__name__})")

    except Exception as e:
        print(f"{attr:<30} = <ERROR>")


print("Layer được chọn:", obj_pick)
layer_name = obj_pick.Layer
print("Layer được chọn:", layer_name)

# ===== tạo selection set =====
name = "SS_" + str(int(time.time()))

# xóa nếu trùng
for s in doc.SelectionSets:
    if s.Name == name:
        s.Delete()

ss = doc.SelectionSets.Add(name)

print("\n👉 Quét đối tượng...")
ss.SelectOnScreen()

# print(f"\nTổng object đã chọn: {ss.Count}")

# ===== lọc theo layer =====
filtered = []

for obj in ss:
    try:
        if obj.Layer == layer_name:
            filtered.append(obj)
    except:
        pass

# ===== kết quả =====
print(f"\n✅ Có {len(filtered)} object thuộc layer {layer_name}")

# ===== in thử =====
for obj in filtered[:10]:
    print(obj.ObjectName, "| Layer:", obj.Layer)



# ===== hàm lấy tọa độ =====
def get_coordinates(obj):
    try:
        # LINE
        if obj.ObjectName == "AcDbLine":
            return {
                "StartPoint": obj.StartPoint,
                "EndPoint": obj.EndPoint
            }

        # CIRCLE
        elif obj.ObjectName == "AcDbCircle":
            return {
                "Center": obj.Center,
                "Radius": obj.Radius
            }

        # ARC
        elif obj.ObjectName == "AcDbArc":
            return {
                "Center": obj.Center,
                "StartAngle": obj.StartAngle,
                "EndAngle": obj.EndAngle
            }

        # TEXT / MTEXT
        elif obj.ObjectName in ["AcDbText", "AcDbMText"]:
            return {
                "InsertionPoint": obj.InsertionPoint
            }

        # POLYLINE
        elif obj.ObjectName in ["AcDbPolyline", "AcDb2dPolyline"]:
            return {
                "Coordinates": obj.Coordinates
            }

        # BLOCK
        elif obj.ObjectName == "AcDbBlockReference":
            return {
                "InsertionPoint": obj.InsertionPoint,
                "Name": obj.Name
            }

        else:
            # fallback nếu có Coordinates
            if hasattr(obj, "Coordinates"):
                return {"Coordinates": obj.Coordinates}

    except:
        pass

    return {"Info": "Không lấy được tọa độ"}

print("\n TỌA ĐỘ OBJECT:")

for i, obj in enumerate(filtered[:10]):  # in 1000 cái đầu
    coords = get_coordinates(obj)
    print(f"\n--- Object {i+1} ---")
    print("Type:", obj.ObjectName)
    print("Layer:", obj.Layer)

    for k, v in coords.items():
        print(f"{k}: {v}")


# ===== xuất Excel =====
wb = Workbook()
ws = wb.active
ws.title = "ToaDo"

# header
ws["A1"] = "X"
ws["B1"] = "Y"


def get_points(obj):
    pts = []

    try:
        t = obj.ObjectName

        if t == "AcDbLine":
            pts.append((obj.StartPoint[0], obj.StartPoint[1]))
            pts.append((obj.EndPoint[0], obj.EndPoint[1]))

        elif t == "AcDbCircle":
            pts.append((obj.Center[0], obj.Center[1]))

        elif t in ["AcDbText", "AcDbMText"]:
            pts.append((obj.InsertionPoint[0], obj.InsertionPoint[1]))

        elif t in ["AcDbPolyline", "AcDb2dPolyline"]:
            coords = obj.Coordinates
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1]))

        elif t == "AcDbBlockReference":
            pts.append((obj.InsertionPoint[0], obj.InsertionPoint[1]))

    except:
        pass

    return pts




row = 2

for obj in filtered:
    pts = get_points(obj)

    for x, y in pts:
        ws[f"A{row}"] = x
        ws[f"B{row}"] = y
        row += 1

# lưu file
file_path = r"C:\Users\KHANGVU\Desktop\toado.xlsx"
wb.save(file_path)

print(f"\n✅ Đã xuất file: {file_path}")


# cleanup
ss.Delete()
